"""サンプルExcelファイルを生成するスクリプト"""
import openpyxl
from openpyxl.drawing.image import Image
from pathlib import Path
from PIL import Image as PILImage
import io

def create_review_checklist(filename, project_name, date, reviewer, approver):
    """レビューチェックリストのサンプルを作成"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "sheet1"

    # サンプルデータを設定
    # E4, E5, E6, N6, M10, M11, M15, M16にデータを入力
    ws['E4'] = project_name
    ws['E5'] = date
    ws['E6'] = reviewer
    ws['N6'] = approver
    ws['M10'] = "設計レビュー"
    ws['M11'] = "OK"
    ws['M15'] = "実装レビュー"
    ws['M16'] = "OK"

    # その他のセルにもダミーデータを追加
    ws['A1'] = "レビューチェックリスト"
    ws['A4'] = "プロジェクト名:"
    ws['A5'] = "日付:"
    ws['A6'] = "担当者:"
    ws['L6'] = "承認者:"

    wb.save(filename)
    print(f"✓ レビューチェックリストを作成: {filename}")

def create_review_record(filename, project_name, reviewer, has_stamp, approval_date="2025-01-15"):
    """レビュー記録表のサンプルを作成（画像付き/なし）"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "sheet1"

    # サンプルデータを設定
    # AE2, AE4, AE5, AE6, AE7, AE8, AB17にデータを入力
    ws['AE2'] = project_name
    ws['AE4'] = "レビュー項目1"
    ws['AE5'] = "レビュー項目2"
    ws['AE6'] = "レビュー項目3"
    ws['AE7'] = f"承認日: {approval_date}"
    ws['AE8'] = f"レビュアー: {reviewer}"
    ws['AB17'] = "備考: 問題なし"

    # その他のセルにもダミーデータを追加
    ws['A1'] = "レビュー記録表"
    ws['AA2'] = "プロジェクト:"

    # BY3に画像を挿入（電子捺印を想定）
    if has_stamp:
        img_data = PILImage.new('RGB', (50, 50), color='red')
        img_buffer = io.BytesIO()
        img_data.save(img_buffer, format='PNG')
        img_buffer.seek(0)

        img = Image(img_buffer)
        img.width = 50
        img.height = 50
        ws.add_image(img, 'BY3')

    wb.save(filename)
    stamp_status = "画像あり" if has_stamp else "画像なし"
    print(f"✓ レビュー記録表を作成: {filename} ({stamp_status})")

def create_other_file():
    """マッチしないファイルも作成（テスト用）"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "sheet1"
    ws['A1'] = "その他のファイル"
    ws['B1'] = "マッチしないファイル"

    wb.save("input/other_document.xlsx")
    print("✓ その他のファイルを作成: input/other_document.xlsx（マッチしないテスト用）")

if __name__ == "__main__":
    print("サンプルExcelファイルを生成中...")

    # サブディレクトリを作成
    Path("input/project_a").mkdir(parents=True, exist_ok=True)
    Path("input/project_b").mkdir(parents=True, exist_ok=True)

    # ===== ペアになるファイル =====

    # プロジェクトA: 完全一致（チェックリスト + 記録表、日付・担当者一致、画像あり）
    create_review_checklist(
        "input/sample_0_レビューチェックリスト_v1.xlsx",
        "プロジェクトA", "2025-01-01", "山田太郎", "承認者A"
    )
    create_review_record(
        "input/レビュー記録表(社内)_2025.xlsx",
        "プロジェクトA", "山田太郎", True, "2025-01-01"
    )

    # プロジェクトB: 完全一致（チェックリスト + 記録表、日付・担当者一致、画像あり）
    create_review_checklist(
        "input/test_0_レビューチェックリスト_v2.xlsx",
        "プロジェクトB", "2025-01-15", "佐藤花子", "承認者B"
    )
    create_review_record(
        "input/レビュー記録表(社外)_202501.xlsx",
        "プロジェクトB", "佐藤花子", True, "2025-01-15"
    )

    # プロジェクトC: 不一致パターン（チェックリスト + 記録表、日付不一致、画像あり）
    create_review_checklist(
        "input/project_a/design_0_レビューチェックリスト_final.xlsx",
        "プロジェクトC", "2025-02-01", "田中次郎", "承認者C"
    )
    create_review_record(
        "input/project_a/レビュー記録表(社員用).xlsx",
        "プロジェクトC", "田中次郎", True, "2025-02-10"  # 日付不一致
    )

    # プロジェクトD: 捺印なしパターン（チェックリスト + 記録表、日付・担当者一致、画像なし）
    create_review_checklist(
        "input/project_b/checklist_0_レビューチェックリスト_draft.xlsx",
        "プロジェクトD（ドラフト）", "2025-01-20", "鈴木一郎", "承認者D"
    )
    create_review_record(
        "input/project_b/レビュー記録表(社内)_draft.xlsx",
        "プロジェクトD（ドラフト）", "鈴木一郎", False, "2025-01-20"
    )

    # ===== ペアのないファイル =====

    # プロジェクトE: チェックリストのみ
    create_review_checklist(
        "input/checklist_0_レビューチェックリスト_solo.xlsx",
        "プロジェクトE（チェックリストのみ）", "2025-02-10", "高橋良子", "承認者E"
    )

    # プロジェクトF: 記録表のみ（画像あり）
    create_review_record(
        "input/レビュー記録表(社内)_solo.xlsx",
        "プロジェクトF（記録表のみ）", "伊藤健太", True, "2025-02-15"
    )

    # マッチしないファイル
    create_other_file()

    print("\n完了！")
    print(f"\n生成されたファイル:")
    print(f"  【ペアあり】")
    print(f"    - プロジェクトA: 完全一致")
    print(f"    - プロジェクトB: 完全一致")
    print(f"    - プロジェクトC: 日付不一致")
    print(f"    - プロジェクトD: 捺印なし")
    print(f"  【ペアなし】")
    print(f"    - プロジェクトE: チェックリストのみ")
    print(f"    - プロジェクトF: 記録表のみ")
    print(f"  合計: 11ファイル（チェックリスト: 5, 記録表: 5, その他: 1）")
